from openpyxl import load_workbook as lw
import unidecode

#Native libraries
import re
import csv
import os
#################

#Reading the Workbook
wb = lw('US-News Ranking.xlsx', data_only= True)

#Getting sheet names
sheets = wb.sheetnames

#Getting rid of previously stored data
files = os.listdir("./")
if "compiled.csv" in files:
    os.remove("compiled.csv")

#Finds which column contains tuition fee
def tuitionColumn(sheetname, wb):
    w_sh = wb[sheetname]
    letters = ["A", "B", "C", "D","E","F","G", "H","I","J"]
    results = []
    for j in range(1, len(letters)):
        cell = w_sh.cell(row=1,column=j).value
        if cell == None:
            continue
        matched = re.search(r"^tuition", cell, re.IGNORECASE)
        if matched != None:
            col = j
            break
        else:
            col = None
            continue
    return col

#Read through University names, Department and Tuition fee
def readSheet(sheetname, wb,tC): #tC = tuition column
    if tC == None:
        return
    i = 2
    w_sh = wb[sheetname]
    with open('compiled.csv','a+',newline='') as f:
        cursor = csv.writer(f)
        while w_sh.cell(row=i, column=tC).value != None:
            university = w_sh.cell(row=i, column=2).value
            university = unidecode.unidecode(university)
            tuition_fee = w_sh.cell(row=i, column=tC).value

            if type(tuition_fee) is float:
                tf = tuition_fee
            else:
                try:
                    tf = float(re.sub(r'[\D]', "", tuition_fee))
                except:
                    tf = ''
            cursor.writerow([university,sheetname,tf])
            i = i + 1

#Main
if __name__ == '__main__':
    lst = []
    for sheet in sheets:
        tc = tuitionColumn(sheet, wb)
        if tc == None:
            continue
        lst.append((sheet, tc))
    for i in lst:
        readSheet(i[0], wb, i[1])
